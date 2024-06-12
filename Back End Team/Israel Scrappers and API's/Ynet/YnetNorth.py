from playwright.sync_api import sync_playwright
from dataclasses import dataclass, asdict, field
import pandas as pd
import os
import re
from openpyxl import Workbook

@dataclass
class Article:
    """Holds article data"""
    title: str
    subtitle: str
    full_text: str
    link: str

@dataclass
class ArticleList:
    """Holds list of Article objects, and saves to both excel and csv"""
    article_list: list[Article] = field(default_factory=list)
    save_at: str = 'output'

    def dataframe(self):
        """Transform article_list to pandas dataframe"""
        return pd.DataFrame([asdict(article) for article in self.article_list])

    def save_to_excel(self, filename):
        """Saves pandas dataframe to excel (xlsx) file"""
        if not os.path.exists(self.save_at):
            os.makedirs(self.save_at)

        filename = re.sub(r'[\\/*?:"<>|]', "_", filename)
        df = self.dataframe()

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active

        # Write the column headers
        headers = df.columns
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=c_idx, value=header)

        # Write the data rows
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook
        wb.save(os.path.join(self.save_at, f"{filename}.xlsx"))

def main():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        page.goto("https://www.ynet.co.il/vacation/category/1174", timeout=60000)
        previously_counted = 0
        total = 1_000_000
        article_list = ArticleList()
        unique_links = set()

        while True:
            page.mouse.wheel(0, 10000)
            page.wait_for_timeout(3000)

            articles = page.locator('//a[contains(@href, "https://www.ynet.co.il/vacation/article")]')

            if articles.count() >= total:
                listings = articles.all()[:total]
                print(f"Total Scraped: {len(listings)}")
                break
            else:
                if articles.count() == previously_counted:
                    listings = articles.all()
                    print(f"Arrived at all available\nTotal Scraped: {len(listings)}")
                    break
                else:
                    previously_counted = articles.count()
                    print(f"Currently Scraped: {previously_counted}")

        for article in articles.all():
            link = article.get_attribute('href')
            if link not in unique_links:
                unique_links.add(link)
                title = article.text_content().strip()

                # Navigate to the article's page to get subtitle and full text
                page.goto(link)
                page.wait_for_timeout(3000)  # Wait for the page to load
                try:
                    main_title = page.locator('//h1[contains(@class, "mainTitle")]').text_content(timeout=10000).strip()
                except:
                    main_title = ""
                try:
                    subtitle = page.locator('//h2[contains(@class, "subTitle")]').text_content(timeout=10000).strip()
                except:
                    subtitle = ""

                try:
                    full_text_elements = page.locator('//span[@data-text="true"]')
                    full_text = "\n".join([element.text_content(timeout=10000).strip() for element in full_text_elements.all()])
                except:
                    full_text = ""

                article_list.article_list.append(Article(title=title, subtitle=subtitle, full_text=full_text, link=link))
                page.go_back()
                page.wait_for_timeout(3000)  # Wait for the main page to load

        article_list.save_to_excel("ynet_articles")

if __name__ == "__main__":
    main()
