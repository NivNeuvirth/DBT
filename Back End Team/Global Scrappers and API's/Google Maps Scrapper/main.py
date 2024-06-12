"""This script serves as an example on how to use Python 
   & Playwright to scrape/extract data from Google Maps"""

from playwright.sync_api import sync_playwright
from dataclasses import dataclass, asdict, field
import pandas as pd
import argparse
import os
import sys
import pyautogui
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



@dataclass
class Business:
    """holds business data"""

    name: str = None
    address: str = None
    website: str = None
    phone_number: str = None
    reviews_count: int = None
    reviews_average: float = None
    latitude: float = None
    longitude: float = None
    
    type_of_business: str = None
    country: str = None
    price: str = None
    hours: str = None
    description: str = None
    #picture_links: str = None


@dataclass
class BusinessList:
    """holds list of Business objects,
    and save to both excel and csv
    """
    business_list: list[Business] = field(default_factory=list)
    save_at = 'output'

    def dataframe(self):
        """transform business_list to pandas dataframe

        Returns: pandas dataframe
        """
        return pd.json_normalize(
            (asdict(business) for business in self.business_list), sep="_"
        )

    def save_to_excel(self, filename):
        """saves pandas dataframe to excel (xlsx) file

        Args:
            filename (str): filename
        """

        if not os.path.exists(self.save_at):
            os.makedirs(self.save_at)
            
        filename = re.sub(r'[\\/*?:"<>|]', "_", filename)

    
    # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
    
    # Write the column headers
        headers = self.dataframe().columns
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=c_idx, value=header)
    
    # Write the data rows
        for r_idx, row in enumerate(self.dataframe().itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):  # Start from index 1 to skip the index column
                ws.cell(row=r_idx, column=c_idx, value=value)
        
    # Save the workbook
        wb.save(f"output/{filename}.xlsx")

    def save_to_csv(self, filename):
        """saves pandas dataframe to csv file

        Args:
            filename (str): filename
        """

        if not os.path.exists(self.save_at):
            os.makedirs(self.save_at)
    # Specify encoding as utf-8
        self.dataframe().to_csv(f"output/{filename}.csv", index=False, encoding="utf-8")

def extract_coordinates_from_url(url: str) -> tuple[float,float]:
    """helper function to extract coordinates from url"""
    
    #coordinates = url.split('3d')[-1].split('/')[0]
    longtitude = url.split('3d')[-1].split('!')[0]
    latitude = url.split('4d')[-1].split('!')[0]
    # return latitude, longitude
    #return float(coordinates.split(',')[0]), float(coordinates.split(',')[1])
    return float(longtitude), float(latitude)



def main():
    
    ########
    # input 
    ########
    
    # read search from arguments
    parser = argparse.ArgumentParser()
    parser.add_argument("-s", "--search", type=str)
    parser.add_argument("-t", "--total", type=int)
    args = parser.parse_args()
    
    if args.search:
        search_list = [args.search]
        
    if args.total:
        total = args.total
    else:
        # if no total is passed, we set the value to random big number
        total = 1_000_000

    if not args.search:
        search_list = []
        # read search from input.txt file
        input_file_name = 'input.txt'
        # Get the absolute path of the file in the current working directory
        input_file_path = os.path.join(os.getcwd(), input_file_name)
        # Check if the file exists
        if os.path.exists(input_file_path):
        # Open the file in read mode
            with open(input_file_path, 'r',encoding='utf-8') as file:
            # Read all lines into a list
                search_list = [line.strip() for line in file.readlines()]#file.readlines()
                
        if len(search_list) == 0:
            print('Error occured: You must either pass the -s search argument, or add searches to input.txt')
            sys.exit()
        
    ###########
    # scraping
    ###########
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        page.goto("https://www.google.com/maps", timeout=60000)
        # wait is added for dev phase. can remove it in production
        #page.wait_for_timeout(5000)
        
        for search_for_index, search_for in enumerate(search_list):
            print(f"-----\n{search_for_index} - {search_for}".strip())

            page.locator('//input[@id="searchboxinput"]').fill(search_for)
            page.wait_for_timeout(3000)
            page.keyboard.press("Enter")
            page.wait_for_timeout(5000)

            # scrolling
            page.hover('//a[contains(@href, "https://www.google.com/maps/place")]')

            # this variable is used to detect if the bot
            # scraped the same number of listings in the previous iteration
            previously_counted = 0
            while True:
                page.mouse.wheel(0, 10000)
                page.wait_for_timeout(3000)

                if (
                    page.locator(
                        '//a[contains(@href, "https://www.google.com/maps/place")]'
                    ).count()
                    >= total
                ):
                    listings = page.locator(
                        '//a[contains(@href, "https://www.google.com/maps/place")]'
                    ).all()[:total]
                    listings = [listing.locator("xpath=..") for listing in listings]
                    print(f"Total Scraped: {len(listings)}")
                    break
                else:
                    # logic to break from loop to not run infinitely
                    # in case arrived at all available listings
                    if (
                        page.locator(
                            '//a[contains(@href, "https://www.google.com/maps/place")]'
                        ).count()
                        == previously_counted
                    ):
                        listings = page.locator(
                            '//a[contains(@href, "https://www.google.com/maps/place")]'
                        ).all()
                        print(f"Arrived at all available\nTotal Scraped: {len(listings)}")
                        break
                    else:
                        previously_counted = page.locator(
                            '//a[contains(@href, "https://www.google.com/maps/place")]'
                        ).count()
                        print(
                            f"Currently Scraped: ",
                            page.locator(
                                '//a[contains(@href, "https://www.google.com/maps/place")]'
                            ).count(),
                        )

            business_list = BusinessList()

            # scraping
            
            i = 0
            for listing in listings:
                try:
                    i = i + 1
                    print(i)
                    listing.click()
                    page.wait_for_timeout(5000)
                    
                    price_xpath = '//div[@class="drwWxc"]'
                    hours_xpath = '//tr[@class="y0skZc"]//td[@class="mxowUb"]/ul/li[@class="G8aQO"]'
                    name_attibute = 'aria-label'
                    type_xpath = '//button[@class="DkEaL " and contains(@jsaction, "category")]'
                    #hours_xpath = '//button[@data-item-id="oh"]//div[contains(@class, "fontBodyMedium")]'
                    address_xpath = '//button[@data-item-id="address"]//div[contains(@class, "fontBodyMedium")]'
                    website_xpath = '//a[@data-item-id="authority"]//div[contains(@class, "fontBodyMedium")]'
                    phone_number_xpath = '//button[contains(@data-item-id, "phone:tel:")]//div[contains(@class, "fontBodyMedium")]'
                    #review_count_xpath = '//button[@jsaction="pane.reviewChart.moreReviews"]//span'
                    review_count_xpath = '//span[contains(@aria-label, "reviews") and contains(text(), "(") and contains(text(), ")")]'
                    reviews_average_xpath = '//div[@jsaction="pane.reviewChart.moreReviews"]//div[@role="img"]'
                    description_xpath = '//div[@class="WeS02d fontBodyMedium"]/div/div[@class="PYvSYb "]'
                    div_xpath = '//div[@class="Uf0tqf loaded"]'

                    
                    
                    
                    
                    
                    business = Business()
                    #if len(listing.get_attribute(type_attribute)) >= 1:
                    #
                    #    business.type_of_business = listing.get_attribute(type_attribute)
                    #else:
                    #    business.type_of_business = ""
                    #
                    if len(listing.get_attribute(name_attibute)) >= 1:
        
                        business.name = listing.get_attribute(name_attibute)
                    else:
                        business.name = ""
                        
                    if page.locator(type_xpath).count() > 0:
                        business.type_of_business = page.locator(type_xpath).all()[0].inner_text()
                    else:
                        business.type_of_business = ""    
                        
                       
                    
                    
                    if page.locator(address_xpath).count() > 0:
                        business.address = page.locator(address_xpath).all()[0].inner_text()
                    else:
                        business.address = ""
                    if page.locator(website_xpath).count() > 0:
                        business.website = page.locator(website_xpath).all()[0].inner_text()
                    else:
                        business.website = ""
                    if page.locator(phone_number_xpath).count() > 0:
                        business.phone_number = page.locator(phone_number_xpath).all()[0].inner_text()
                    else:
                        business.phone_number = ""
                    if page.locator(price_xpath).count() > 0:
                        price_elements = page.locator(price_xpath).all()
                        if len(price_elements) > 0:
                            prices = [price_element.inner_text() for price_element in price_elements]
                            business.price = ", ".join(prices)
                    else:
                        business.price = ""
                    
                    address_parts = business.address.split(', ')
                    business.country = address_parts[-1]
                    
                    
                    
                    if page.locator(hours_xpath).count() > 0:
                        hours_elements = page.locator(hours_xpath).all()
                        hours_list = [hour_element.inner_text() for hour_element in hours_elements]
                        business.hours = str(", ".join(hours_list))
                        
                    else:
                        business.hours = ""
                    
                    
                    

                    if page.locator(description_xpath).count() > 0:
                        business.description = page.locator(description_xpath).inner_text()
                    
                    if page.locator(review_count_xpath).count() > 0:
                    # Get the inner text of the element
                        inner_text = page.locator(review_count_xpath).inner_text()
    
                        # Extract the text within the parentheses
                        review_text = inner_text.split('(')[-1].split(')')[0]
    
                    # Remove commas and whitespace, then convert to integer
                        reviews_count = int(review_text.replace(',', '').strip())
    
                        business.reviews_count = reviews_count
                    else:
                        business.reviews_count = ""
                        
                        
               
                        
                    if page.locator(reviews_average_xpath).count() > 0:
                        review_average_text = page.locator(reviews_average_xpath).get_attribute(name_attibute)
                        review_average_cleaned = re.sub(r'[^\d.]', '', review_average_text)
                        business.reviews_average = float(review_average_cleaned)
                    else:
                        business.reviews_average = ""
                    
                    
                    business.latitude, business.longitude = extract_coordinates_from_url(page.url)

                    business_list.business_list.append(business)
                except Exception as e:
                    print(f'Error occured: {e}')
            
            #########
            # output
            #########
            #business_list.save_to_csv(f"google_maps_data_{search_for}".replace(' ', '_'))
            business_list.save_to_excel(f"google_maps_data_{search_for}".replace(' ', '_'))

        browser.close()


if __name__ == "__main__":
    main()
